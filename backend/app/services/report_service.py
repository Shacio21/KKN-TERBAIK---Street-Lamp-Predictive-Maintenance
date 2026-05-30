"""Report generation service.

Generates Excel (openpyxl) and PDF (reportlab) reports for:
- lamp_status: Lamp health, status, battery, risk overview
- energy: Solar generation, consumption, cost savings
- maintenance: Ticket history, response times
- predictive: ML prediction results, risk distribution
"""

import io
import logging
from datetime import date, datetime
from pathlib import Path

from sqlalchemy import func, select
from sqlalchemy.ext.asyncio import AsyncSession

from app.models import (
    Alert, EnergyCostConfig, Lamp, LampStatus, Place, Prediction,
    RepairTicket, Telemetry,
)

logger = logging.getLogger(__name__)

REPORT_DIR = Path(__file__).parent.parent.parent / "reports"


async def generate_lamp_status_report(
    db: AsyncSession, start_date: date, end_date: date, fmt: str = "csv"
) -> io.BytesIO:
    """Generate lamp status report."""
    lamps = list(
        await db.scalars(
            select(Lamp).where(Lamp.is_deleted.is_(False)).order_by(Lamp.lamp_code)
        )
    )

    headers = [
        "Kode Lampu", "Status", "Health Score", "Risk Level", "Battery %",
        "Brightness %", "Signal %", "Model", "Power (W)", "Firmware",
        "Terakhir Online", "Telemetry Count", "ML Ready",
    ]

    rows = []
    for lamp in lamps:
        rows.append([
            lamp.lamp_code,
            lamp.status or "unknown",
            f"{float(lamp.health_score):.1f}" if lamp.health_score is not None else "—",
            lamp.risk_level or "unknown",
            f"{float(lamp.last_battery_level):.0f}" if lamp.last_battery_level is not None else "—",
            f"{float(lamp.last_brightness):.0f}" if lamp.last_brightness is not None else "—",
            f"{float(lamp.last_signal_strength):.0f}" if lamp.last_signal_strength is not None else "—",
            lamp.model or "—",
            f"{float(lamp.power_rating):.0f}" if lamp.power_rating is not None else "—",
            lamp.firmware_version or "—",
            lamp.last_seen.isoformat() if lamp.last_seen else "—",
            lamp.telemetry_count or 0,
            "Ya" if lamp.ml_ready else "Tidak",
        ])

    # Summary row
    total = len(lamps)
    online = sum(1 for l in lamps if l.status == LampStatus.online.value)
    fault = sum(1 for l in lamps if l.status == LampStatus.fault.value)

    if fmt == "csv":
        return _build_csv(headers, rows, f"Laporan Status Lampu ({start_date} s/d {end_date})")
    else:
        return _build_excel(
            headers, rows,
            title=f"Laporan Status Lampu",
            subtitle=f"Periode: {start_date} s/d {end_date}",
            summary={
                "Total Lampu": total, "Online": online, "Fault": fault,
                "Offline": sum(1 for l in lamps if l.status == LampStatus.offline.value),
            },
        )


async def generate_energy_report(
    db: AsyncSession, start_date: date, end_date: date, fmt: str = "csv"
) -> io.BytesIO:
    """Generate energy & sustainability report."""
    from datetime import timezone

    start_dt = datetime.combine(start_date, datetime.min.time()).replace(tzinfo=timezone.utc)
    end_dt = datetime.combine(end_date, datetime.max.time()).replace(tzinfo=timezone.utc)

    # Aggregate telemetry per lamp
    result = await db.execute(
        select(
            Telemetry.lamp_id,
            func.avg(Telemetry.solar_power).label("avg_solar"),
            func.avg(Telemetry.power).label("avg_consumed"),
            func.sum(Telemetry.solar_energy_today).label("total_solar"),
            func.count().label("readings"),
        )
        .where(Telemetry.time.between(start_dt, end_dt))
        .group_by(Telemetry.lamp_id)
    )
    energy_data = result.all()

    # Get lamp codes
    lamp_map = {}
    if energy_data:
        lamp_ids = [r[0] for r in energy_data]
        lamps = list(await db.scalars(select(Lamp).where(Lamp.id.in_(lamp_ids))))
        lamp_map = {l.id: l for l in lamps}

    headers = [
        "Kode Lampu", "Avg Solar (W)", "Avg Konsumsi (W)",
        "Total Solar (Wh)", "Readings", "Efisiensi (%)",
    ]
    rows = []
    for row in energy_data:
        lamp = lamp_map.get(row[0])
        avg_solar = float(row[1] or 0)
        avg_consumed = float(row[2] or 0)
        efficiency = (avg_solar / avg_consumed * 100) if avg_consumed > 0 else 0
        rows.append([
            lamp.lamp_code if lamp else str(row[0]),
            f"{avg_solar:.2f}",
            f"{avg_consumed:.2f}",
            f"{float(row[3] or 0):.2f}",
            row[4],
            f"{efficiency:.1f}",
        ])

    if fmt == "csv":
        return _build_csv(headers, rows, f"Laporan Energi ({start_date} s/d {end_date})")
    else:
        return _build_excel(
            headers, rows,
            title="Laporan Energi & Sustainability",
            subtitle=f"Periode: {start_date} s/d {end_date}",
        )


async def generate_maintenance_report(
    db: AsyncSession, start_date: date, end_date: date, fmt: str = "csv"
) -> io.BytesIO:
    """Generate maintenance ticket report."""
    from datetime import timezone

    start_dt = datetime.combine(start_date, datetime.min.time()).replace(tzinfo=timezone.utc)
    end_dt = datetime.combine(end_date, datetime.max.time()).replace(tzinfo=timezone.utc)

    tickets = list(
        await db.scalars(
            select(RepairTicket)
            .where(RepairTicket.created_at.between(start_dt, end_dt))
            .order_by(RepairTicket.created_at.desc())
        )
    )

    # Get lamp codes
    lamp_ids = list({t.lamp_id for t in tickets})
    lamp_map = {}
    if lamp_ids:
        lamps = list(await db.scalars(select(Lamp).where(Lamp.id.in_(lamp_ids))))
        lamp_map = {l.id: l for l in lamps}

    headers = [
        "Tiket", "Lampu", "Prioritas", "Status", "Dibuat", "Selesai",
        "Durasi Resolusi (jam)", "Catatan Resolusi",
    ]
    rows = []
    for t in tickets:
        lamp = lamp_map.get(t.lamp_id)
        duration = None
        if t.resolved_at and t.created_at:
            duration = (t.resolved_at - t.created_at).total_seconds() / 3600
        rows.append([
            t.title,
            lamp.lamp_code if lamp else "—",
            t.priority,
            t.status,
            t.created_at.isoformat() if t.created_at else "—",
            t.resolved_at.isoformat() if t.resolved_at else "—",
            f"{duration:.1f}" if duration else "—",
            t.resolution_note or "—",
        ])

    if fmt == "csv":
        return _build_csv(headers, rows, f"Laporan Maintenance ({start_date} s/d {end_date})")
    else:
        total = len(tickets)
        resolved = sum(1 for t in tickets if t.status == "resolved")
        return _build_excel(
            headers, rows,
            title="Laporan Pemeliharaan",
            subtitle=f"Periode: {start_date} s/d {end_date}",
            summary={"Total Tiket": total, "Resolved": resolved, "Open": total - resolved},
        )


async def generate_predictive_report(
    db: AsyncSession, start_date: date, end_date: date, fmt: str = "csv"
) -> io.BytesIO:
    """Generate predictive maintenance (ML) report."""
    from datetime import timezone

    start_dt = datetime.combine(start_date, datetime.min.time()).replace(tzinfo=timezone.utc)
    end_dt = datetime.combine(end_date, datetime.max.time()).replace(tzinfo=timezone.utc)

    preds = list(
        await db.scalars(
            select(Prediction)
            .where(Prediction.predicted_at.between(start_dt, end_dt))
            .order_by(Prediction.predicted_at.desc())
        )
    )

    lamp_ids = list({p.lamp_id for p in preds if p.lamp_id})
    lamp_map = {}
    if lamp_ids:
        lamps = list(await db.scalars(select(Lamp).where(Lamp.id.in_(lamp_ids))))
        lamp_map = {l.id: l for l in lamps}

    headers = [
        "Lampu", "Risk Level", "Failure Prob. (%)", "Hari hingga Gagal",
        "Confidence (%)", "Model Version", "Rekomendasi", "Tanggal Prediksi",
    ]
    rows = []
    for p in preds:
        lamp = lamp_map.get(p.lamp_id)
        rows.append([
            lamp.lamp_code if lamp else "—",
            p.risk_level or "—",
            f"{float(p.failure_probability)*100:.1f}" if p.failure_probability is not None else "—",
            p.days_to_failure if p.days_to_failure is not None else "—",
            f"{float(p.confidence)*100:.1f}" if p.confidence is not None else "—",
            p.model_version or "—",
            p.recommendation or "—",
            p.predicted_at.isoformat() if p.predicted_at else "—",
        ])

    if fmt == "csv":
        return _build_csv(headers, rows, f"Laporan Prediktif ({start_date} s/d {end_date})")
    else:
        high = sum(1 for p in preds if p.risk_level == "high")
        medium = sum(1 for p in preds if p.risk_level == "medium")
        low = sum(1 for p in preds if p.risk_level == "low")
        return _build_excel(
            headers, rows,
            title="Laporan Prediksi Maintenance (ML)",
            subtitle=f"Periode: {start_date} s/d {end_date}",
            summary={"Total Prediksi": len(preds), "High Risk": high, "Medium": medium, "Low": low},
        )


# ──────────────────────────────────────────────────
# Output builders
# ──────────────────────────────────────────────────

def _build_csv(headers: list, rows: list, title: str = "") -> io.BytesIO:
    """Build CSV output."""
    import csv

    buffer = io.BytesIO()
    wrapper = io.TextIOWrapper(buffer, encoding="utf-8-sig", newline="")  # BOM for Excel

    writer = csv.writer(wrapper)
    if title:
        writer.writerow([title])
        writer.writerow([])
    writer.writerow(headers)
    for row in rows:
        writer.writerow(row)

    wrapper.flush()
    wrapper.detach()
    buffer.seek(0)
    return buffer


def _build_excel(
    headers: list, rows: list, title: str = "", subtitle: str = "",
    summary: dict | None = None,
) -> io.BytesIO:
    """Build Excel output with openpyxl."""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "Laporan"

    # Header styles
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
    title_font = Font(bold=True, size=14, color="00D4FF")
    subtitle_font = Font(size=10, color="94A3B8")

    row_idx = 1
    # Title
    ws.cell(row=row_idx, column=1, value=title).font = title_font
    row_idx += 1
    if subtitle:
        ws.cell(row=row_idx, column=1, value=subtitle).font = subtitle_font
        row_idx += 1

    # Summary
    if summary:
        row_idx += 1
        for i, (key, val) in enumerate(summary.items()):
            col = (i * 2) + 1
            ws.cell(row=row_idx, column=col, value=key).font = Font(bold=True, size=10)
            ws.cell(row=row_idx, column=col + 1, value=val).font = Font(size=10, color="00D4FF")
        row_idx += 1

    row_idx += 1

    # Headers
    for col_i, header in enumerate(headers, 1):
        cell = ws.cell(row=row_idx, column=col_i, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
    row_idx += 1

    # Data rows
    for row_data in rows:
        for col_i, value in enumerate(row_data, 1):
            ws.cell(row=row_idx, column=col_i, value=value)
        row_idx += 1

    # Auto-width columns
    for col_i in range(1, len(headers) + 1):
        max_len = max(
            len(str(ws.cell(row=r, column=col_i).value or ""))
            for r in range(1, row_idx)
        )
        ws.column_dimensions[get_column_letter(col_i)].width = min(max_len + 4, 40)

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


# ──────────────────────────────────────────────────
# Main dispatcher
# ──────────────────────────────────────────────────

GENERATORS = {
    "lamp_status": generate_lamp_status_report,
    "energy": generate_energy_report,
    "maintenance": generate_maintenance_report,
    "predictive": generate_predictive_report,
}


async def generate_report(
    db: AsyncSession, report_type: str, start_date: date, end_date: date, fmt: str = "csv"
) -> io.BytesIO:
    """Generate a report by type."""
    generator = GENERATORS.get(report_type)
    if not generator:
        raise ValueError(f"Unknown report type: {report_type}")
    return await generator(db, start_date, end_date, fmt)
